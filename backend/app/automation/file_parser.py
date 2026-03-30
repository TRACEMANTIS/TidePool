"""File parsing utilities for contact import.

Supports Excel (.xlsx) and CSV files.  Excel files are opened in
read_only=True mode so that openpyxl streams rows without loading the
entire workbook into memory.  CSV encoding is auto-detected via the
chardet library when available, falling back to utf-8.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Generator

# openpyxl is imported lazily inside parse_excel so the module can still
# be imported in environments that only process CSV files.

# ---------------------------------------------------------------------------
# Column-mapping type
# ---------------------------------------------------------------------------

ColumnMapping = dict[str, str | None]
"""Maps logical field names to source column headers.

Expected keys: ``email``, ``first_name``, ``last_name``, ``department``.
A value of ``None`` means the field is not mapped.
"""

# Email validation -- intentionally permissive; we only reject obvious junk.
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Heuristics for auto-detecting column purpose from header text.
_COLUMN_HINTS: dict[str, list[str]] = {
    "email": ["email", "e-mail", "mail", "email_address", "emailaddress"],
    "first_name": ["first_name", "firstname", "first", "given_name", "givenname"],
    "last_name": ["last_name", "lastname", "last", "surname", "family_name"],
    "department": ["department", "dept", "division", "group", "team"],
}

BATCH_SIZE = 1000


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalise_header(header: str) -> str:
    """Lower-case and strip whitespace from a header string."""
    return header.strip().lower().replace(" ", "_")


def _detect_csv_encoding(file_path: str | Path) -> str:
    """Best-effort encoding detection for a CSV file."""
    try:
        import chardet  # type: ignore[import-untyped]
    except ImportError:
        return "utf-8"

    raw = Path(file_path).read_bytes(8192) if hasattr(Path, "read_bytes") else b""
    # read_bytes does not support a length argument -- read the whole file
    # but cap at 32 KB to keep detection fast.
    raw = Path(file_path).read_bytes()[:32768]
    result = chardet.detect(raw)
    return result.get("encoding") or "utf-8"


def _map_row(row: dict[str, Any], column_mapping: ColumnMapping) -> dict[str, Any]:
    """Extract mapped fields from a raw row dict, returning a contact dict."""
    contact: dict[str, Any] = {}

    email_col = column_mapping.get("email")
    if email_col and email_col in row:
        contact["email"] = str(row[email_col]).strip()
    else:
        contact["email"] = ""

    for field in ("first_name", "last_name", "department"):
        col = column_mapping.get(field)
        if col and col in row:
            val = row[col]
            contact[field] = str(val).strip() if val is not None else None
        else:
            contact[field] = None

    return contact


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel(
    file_path: str | Path,
    column_mapping: ColumnMapping,
) -> Generator[dict[str, Any], None, None]:
    """Yield contact dicts from an Excel (.xlsx) workbook.

    Uses ``openpyxl`` in ``read_only=True`` mode for memory-efficient
    streaming of large files.
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return

        headers: list[str] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell is not None else "" for cell in row]
                continue

            # Skip completely blank rows.
            if all(cell is None for cell in row):
                continue

            raw = dict(zip(headers, row))
            yield _map_row(raw, column_mapping)
    finally:
        wb.close()


def parse_csv(
    file_path: str | Path,
    column_mapping: ColumnMapping,
) -> Generator[dict[str, Any], None, None]:
    """Yield contact dicts from a CSV file.

    Encoding is auto-detected when chardet is available.
    """
    encoding = _detect_csv_encoding(file_path)

    with open(file_path, newline="", encoding=encoding, errors="replace") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            yield _map_row(row, column_mapping)


def detect_columns(file_path: str | Path) -> list[dict[str, Any]]:
    """Detect columns and sample values from the first rows of a file.

    Returns a list of dicts, each with keys ``name``, ``sample_values``
    (up to 5), and ``suggested_mapping`` (one of the logical field names
    or ``None``).
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    headers: list[str] = []
    sample_rows: list[dict[str, Any]] = []

    if suffix in (".xlsx", ".xls"):
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(c) if c is not None else "" for c in row]
                elif row_idx <= 5:
                    sample_rows.append(dict(zip(headers, row)))
                else:
                    break
        finally:
            wb.close()
    elif suffix == ".csv":
        encoding = _detect_csv_encoding(path)
        with open(path, newline="", encoding=encoding, errors="replace") as fh:
            reader = csv.DictReader(fh)
            headers = list(reader.fieldnames or [])
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                sample_rows.append(row)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    results: list[dict[str, Any]] = []
    for header in headers:
        norm = _normalise_header(header)
        suggested: str | None = None
        for field, hints in _COLUMN_HINTS.items():
            if norm in hints:
                suggested = field
                break

        samples = [
            str(row.get(header, ""))
            for row in sample_rows
            if row.get(header) is not None
        ]
        results.append({
            "name": header,
            "sample_values": samples[:5],
            "suggested_mapping": suggested,
        })

    return results


def validate_emails(
    contacts: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Split a contact list into (valid, invalid) based on email format.

    Validation is intentionally permissive -- we only reject entries with
    obviously malformed or missing addresses.
    """
    valid: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []

    for contact in contacts:
        email = contact.get("email", "")
        if email and _EMAIL_RE.match(email):
            valid.append(contact)
        else:
            invalid.append(contact)

    return valid, invalid
