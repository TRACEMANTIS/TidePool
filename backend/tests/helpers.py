"""Test helper utilities for generating test data."""

import csv
import io
import random
import string
import uuid
from typing import Any

from httpx import AsyncClient


def random_email() -> str:
    """Generate a random email address."""
    local = "".join(random.choices(string.ascii_lowercase, k=8))
    domain = "".join(random.choices(string.ascii_lowercase, k=6))
    return f"{local}@{domain}.test"


def random_string(length: int = 12) -> str:
    """Generate a random alphanumeric string."""
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def random_campaign_data() -> dict[str, Any]:
    """Generate a dict of valid campaign creation data."""
    return {
        "name": f"Test Campaign {uuid.uuid4().hex[:8]}",
        "description": "Automated test campaign",
        "template_id": 1,
        "smtp_profile_id": 1,
        "addressbook_id": 1,
        "landing_page_id": None,
        "scheduled_at": None,
    }


def create_test_csv(rows: int, columns: list[str] | None = None) -> bytes:
    """Generate a CSV file in memory with the specified rows and columns.

    Args:
        rows: Number of data rows to generate.
        columns: Column header names. Defaults to standard contact columns.

    Returns:
        UTF-8 encoded bytes of the CSV content.
    """
    if columns is None:
        columns = ["email", "first_name", "last_name", "department"]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)

    for i in range(rows):
        row = []
        for col in columns:
            if col == "email":
                row.append(f"user{i}@example-{uuid.uuid4().hex[:6]}.test")
            elif col == "first_name":
                row.append(f"First{i}")
            elif col == "last_name":
                row.append(f"Last{i}")
            elif col == "department":
                departments = ["Engineering", "Sales", "HR", "Finance", "Marketing"]
                row.append(departments[i % len(departments)])
            else:
                row.append(f"value_{col}_{i}")
        writer.writerow(row)

    return buf.getvalue().encode("utf-8")


def create_test_xlsx(rows: int, columns: list[str] | None = None) -> bytes:
    """Generate an Excel (.xlsx) file in memory with the specified rows and columns.

    Args:
        rows: Number of data rows to generate.
        columns: Column header names. Defaults to standard contact columns.

    Returns:
        Bytes of the XLSX file content.
    """
    from openpyxl import Workbook

    if columns is None:
        columns = ["email", "first_name", "last_name", "department"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Write header row.
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows.
    for i in range(rows):
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name == "email":
                value = f"user{i}@example-{uuid.uuid4().hex[:6]}.test"
            elif col_name == "first_name":
                value = f"First{i}"
            elif col_name == "last_name":
                value = f"Last{i}"
            elif col_name == "department":
                departments = ["Engineering", "Sales", "HR", "Finance", "Marketing"]
                value = departments[i % len(departments)]
            else:
                value = f"value_{col_name}_{i}"
            ws.cell(row=i + 2, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def create_test_user(
    client: AsyncClient,
    admin: bool = False,
    auth_headers: dict | None = None,
) -> dict[str, Any]:
    """Register a test user and return a dict with user data and auth token.

    Args:
        client: The async test client.
        admin: Whether to create an admin user.
        auth_headers: Headers with admin credentials (required to register users).

    Returns:
        Dict with keys: id, username, email, token, headers.
    """
    username = f"testuser_{uuid.uuid4().hex[:8]}"
    email = f"{username}@test.local"
    password = "TestP@ssw0rd!2026"

    register_payload = {
        "username": username,
        "email": email,
        "password": password,
        "is_admin": admin,
    }

    resp = await client.post(
        "/api/v1/auth/register",
        json=register_payload,
        headers=auth_headers or {},
    )
    user_data = resp.json()

    # Log in to get a token.
    login_resp = await client.post(
        "/api/v1/auth/login",
        json={"username": username, "password": password},
    )
    tokens = login_resp.json()

    return {
        "id": user_data.get("id"),
        "username": username,
        "email": email,
        "password": password,
        "token": tokens.get("access_token"),
        "refresh_token": tokens.get("refresh_token"),
        "headers": {"Authorization": f"Bearer {tokens.get('access_token')}"},
    }
